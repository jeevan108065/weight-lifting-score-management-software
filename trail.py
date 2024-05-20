from tkinter import *
from openpyxl import load_workbook
from openpyxl import *




print()
def add():
    workbook = Workbook()
    try:
        workbook = load_workbook(filename="sports_test.xlsx")
    except:
        print("no file")
    sheet = workbook.active
    i=0
    for j in range(3,17):
        if sheet.cell(j, 2).value==53:
            i=j
            break
    cell1="J"+str(i)
    sheet[cell1]=next_weight_Entry.get()
    workbook.save(filename="sports_test.xlsx")
    sort()
def sort():
    import pandas as pd
    xl = pd.ExcelFile("sports_test.xlsx")
    df = xl.parse("Sheet")
    df = df.sort_values(by=["category","trail1","trail2","trail3"])
    writer = pd.ExcelWriter('sports_test_sorted.xlsx')
    df.to_excel(writer,sheet_name='Sheet',
                columns=['chest_no','name','university','category','trail1','trail2','trail3','next_weight'],
                index=False)
    writer.save()
next_weight_win=Tk()
next_weight_win.geometry("100x100")
next_weight_win.configure(bg='black')
next_weight_Entry=Entry(next_weight_win,width=3)
next_weight_Entry.pack()
add_btn=Button(next_weight_win,text="ADD",command=add)
add_btn.pack()
next_weight_win.mainloop()