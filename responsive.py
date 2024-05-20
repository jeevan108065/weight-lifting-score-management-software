import time
import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import ttk, filedialog
import numpy
import pandas as pd
import pyautogui
from PIL import ImageTk, Image
from openpyxl import load_workbook
from openpyxl import *
from playsound import playsound

n=3
width, height= pyautogui.size()
geometry=str(width)+"x"+str(height)

# creating Tk window
root = Tk()

# setting geometry of tk window
root.geometry(geometry)
root.configure(bg='black')
# Using title() to display a message in
# the dialogue box of the message in the
# title bar.
root.title("WEIGHT LIFTING")


#logo placing
img =Image.open('logopng.png')
bg = ImageTk.PhotoImage(img)
bg_image= Label(root, image=bg,bg='black')
bg_image.place(anchor=tk.NW,relx=0.01,rely=0.01)
#bg_image.place(x=0,y=0)
img2 =Image.open('AIU_logo.png')
bg2 = ImageTk.PhotoImage(img2)
bg_image2= Label(root, image=bg2,bg='black')
bg_image2.place(anchor=tk.NE,relx=0.98,rely=0.02)
competition_name_label=Label(root,text="All India Inter-Univesity \n Weightlifting Women Championship 2021-22",
                             anchor="c",fg="red",bg='black',font=("Comic Sans MS",int(height/25),"bold"))
competition_name_label.place(anchor=tk.W, relx=0.15,rely=0.1)
#logo ends

#openinng players data

workbook = Workbook()
try:
    workbook = load_workbook(filename="sports_test.xlsx")
except:
    print("no file")
global sheet
sheet = workbook.active


#timer code
# Declaration of variables
hour=StringVar()
minute=StringVar()
second=StringVar()
# setting the default value as 0
hour.set("00")
minute.set("00")
second.set("60")
# Use of Entry class to take input from the user
hourEntry= Entry(root, width=3, font=("Arial",58,""),
				textvariable=hour)
minuteEntry= Entry(root, width=3, font=("Arial",58,""),
				textvariable=minute)
#minuteEntry.place(x=(width/2)-58,y=20)
secondEntry= Entry(root, width=2,fg="white",bg='black', font=("Arial",80,""),textvariable=second)
secondEntry.place(x=(width/2),y=height-(height/4))
Countdown_label=Label(root,text="COUNT DOWN:",anchor="e",font=("Comic Sans MS",int(height/29)))
Countdown_label.place(anchor=tk.SW, relx=0.30,y=665)

# the input provided by the user is
# stored in here :temp
global temp
global stop

temp = int(second.get())
def submit(temp,stop=False):
    #temp = int(second.get())
    # divmod(firstvalue = temp//60, secondvalue = temp%60)
    mins,secs = divmod(temp,60)
	# Converting the input entered in mins or secs to hours,
	# mins ,secs(input = 110 min --> 120*60 = 6600 => 1hr :
	# 50min: 0sec)
    hours=0
    if mins >60:		
		# divmod(firstvalue = temp//60, secondvalue
		# = temp%60)
        hours, mins = divmod(mins, 60)		
	# using format () method to store the value up to
	# two decimal places
    hour.set("{0:2d}".format(hours))
    minute.set("{0:2d}".format(mins))
    second.set("{0:2d}".format(secs))
	# updating the GUI window after decrementing the
	# temp value every time
    root.update()
    time.sleep(1)
    # when temp value = 0; then a messagebox pop's up
	# with a message:"Time's up"
    if (temp<30):secondEntry['fg']="red"
    if (temp==30):playsound("beep-1.mp3")
    if (temp == 0):
        playsound("beep-2.mp3")
        messagebox.showinfo("Time Countdown", "Time's up ")
    # after every one sec the value of temp will be decremented
	# by one
    if stop==False:
        temp -= 1
    if temp >-1:
        submit(temp,stop)


def next_person():
    second.set("60")
    global n
    save(n,trail1_Entry.get(),trail2_Entry.get(),trail3_Entry.get())

    xl = pd.ExcelFile("sports_test.xlsx")
    df = xl.parse("Sheet")
    df = df.sort_values(by="category")
    print(df)
    writer = pd.ExcelWriter('sports_test_sorted.xlsx')
    df.to_excel(writer,sheet_name='Sheet',columns=['chest_no','name','university','category','trail1','trail2','trail3'],index=False)
    writer.save()
    
    n+=1
    status_display['text'] = sheet.cell(row=n, column=2).value
    player_label['text'] =sheet.cell(row=n, column=3).value.upper()
    university_label['text'] =sheet.cell(row=n, column=4).value.upper()
    category_label['text'] ="CATEGORY:"+sheet.cell(row=n, column=5).value.upper()
    trail1.set(sheet.cell(row=n, column=6).value)
    trail2.set(sheet.cell(row=n, column=7).value)
    trail3.set(sheet.cell(row=n, column=8).value)
    secondEntry['fg']='white'
    if trail1.get()=="?":trail1_Entry['bg']="yellow"
    elif trail1.get()=="-":trail1_Entry['bg']="red"
    else:trail1_Entry['bg']="green"
    if trail2.get()=="?":trail2_Entry['bg']="yellow"
    elif trail2.get()=="-":trail2_Entry['bg']="red"
    else:trail2_Entry['bg']="green"
    if trail3.get()=="?":trail3_Entry['bg']="yellow"
    elif trail3.get()=="-":trail3_Entry['bg']="red"
    else:trail3_Entry['bg']="green"
    submit(temp)
    root.update()
    print(n)
	
		
#timer code ends
next_btn=Button(root,text="SAVE & NEXT",command=next_person)
next_btn.place(x=(width-2*int(width/14)),y=height-2*int(height/10))
#jersey display
chest_no=sheet.cell(row=n, column=2).value
chest_label=Label(root,text="TRUNK NO:",anchor="e",fg="white",bg='black',font=("Comic Sans MS",int(height/31)))
chest_label.place(anchor=tk.NE, relx=0.15,rely=0.33)
status_display= Label(root, text=chest_no,bd=3, relief=SUNKEN, anchor="e" ,font=("Comic Sans MS",200))
status_display.place(x=10,y=4*height/10)
#jersey display end

#player name
player_name=sheet.cell(row=n, column=3).value.upper()
player_label=Label(root,text=player_name,fg="white",bg='black', relief=RAISED,anchor="c",font=("Comic Sans MS",int(height/24)))
player_label.place(anchor=tk.N, relx=0.53,rely=0.22)
#player name end

#university name
university_name=sheet.cell(row=n, column=4).value.upper()
university_label=Label(root,text=university_name,fg="white",bg='black',anchor="e",font=("Comic Sans MS",int(height/28)))
university_label.place(anchor=tk.N, relx=0.53,rely=0.31)
#university name ends

#category
category_name=sheet.cell(row=n, column=5).value.upper()
category_label=Label(root,text="CATEGORY:"+category_name,fg="yellow", bg='black',relief=RAISED,anchor="e",font=("Comic Sans MS",int(height/25)))
category_label.place(anchor=tk.NW,relx=0.01,rely=0.23)
#category end

#event name
event_name="SNATCH"
event_label=Label(root,text=event_name,fg="white",bg='black', relief=SUNKEN,anchor="e",font=("Comic Sans MS",int(height/28)))
event_label.place(anchor=tk.S, relx=0.53,rely=0.5)
#event name end

#score card
trail1=StringVar()
trail2=StringVar()
trail3=StringVar()

trail1_Entry= Entry(root, width=3, font=("Arial",58,""),textvariable=trail1)
trail1_Entry.insert(0,sheet.cell(row=n, column=6).value)
if trail1.get()=="?":trail1_Entry['bg']="yellow"
elif trail1.get()=="-":trail1_Entry['bg']="red"
else:trail1_Entry['bg']="green"
trail2_Entry= Entry(root, width=3, font=("Arial",58,""),textvariable=trail2)
trail2_Entry.insert(0,sheet.cell(row=n, column=7).value)
if trail2.get()=="?":trail2_Entry['bg']="yellow"
elif trail2.get()=="-":trail2_Entry['bg']="red"
else:trail2_Entry['bg']="green"
trail3_Entry= Entry(root, width=3, font=("Arial",58,""),textvariable=trail3)
trail3_Entry.insert(0,sheet.cell(row=n, column=8).value)
if trail3.get()=="?":trail3_Entry['bg']="yellow"
elif trail3.get()=="-":trail3_Entry['bg']="red"
else:trail3_Entry['bg']="green"

trail1_Entry.place(anchor=tk.SE, relx=0.4685,rely=0.75, height=165)
trail2_Entry.place(anchor=tk.S, relx=0.52,rely=0.75, height=165)
trail3_Entry.place(anchor=tk.SW, relx=0.572,rely=0.75, height=165)
#score card end

#adding data to excel
cell1="F"+str(n)
cell2="G"+str(n)
cell3="H"+str(n)
sheet[cell1]=trail1_Entry.get()
sheet[cell2]=trail2_Entry.get()
sheet[cell3]=trail3_Entry.get()
def save(n,t1,t2,t3):
    workbook = Workbook()
    try:
        workbook = load_workbook(filename="sports_test.xlsx")
    except:
        print("no file")
    sheet = workbook.active
    cell1="F"+str(n)
    cell2="G"+str(n)
    cell3="H"+str(n)
    sheet[cell1]=trail1_Entry.get()
    sheet[cell2]=trail2_Entry.get()
    sheet[cell3]=trail3_Entry.get()
    workbook.save(filename="sports_test.xlsx")
save(n,trail1_Entry.get(),trail2_Entry.get(),trail3_Entry.get())
def pass1():
    if trail1_Entry.get()=="?" and trail1_Entry.get()!="-":
        trail1_Entry.delete(0, 'end')
        trail1_Entry.insert(0,current_weight_Entry.get())
        trail1_Entry['bg']="green"
    elif trail2_Entry.get()=="?" and trail2_Entry.get()!="-":
        trail2_Entry.delete(0, 'end')
        trail2_Entry.insert(0,current_weight_Entry.get())
        trail2_Entry['bg']="green"
    elif trail3_Entry.get()=="?" and trail3_Entry.get()!="-":
        trail3_Entry.delete(0, 'end')
        trail3_Entry.insert(0,current_weight_Entry.get())
        trail3_Entry['bg']="green"
    stop=True
    playsound("beep-1.mp3")
    submit(temp,stop)
    root.update()
    pass
def fail():
    if trail1_Entry.get()=="?" and trail1_Entry.get()!="-":
        trail1_Entry.delete(0, 'end')
        trail1_Entry.insert(0,"-")
        trail1_Entry['bg']="red"
    elif trail2_Entry.get()=="?" and trail2_Entry.get()!="-":
        trail2_Entry.delete(0, 'end')
        trail2_Entry.insert(0,"-")
        trail2_Entry['bg']="red"
    elif trail3_Entry.get()=="?" and trail3_Entry.get()!="-":
        trail3_Entry.delete(0, 'end')
        trail3_Entry.insert(0,"-")
        trail3_Entry['bg']="red"
    stop=True
    playsound("beep-1.mp3")
    submit(temp,stop)
    root.update()
    pass
#current weight
player_name2=sheet.cell(row=n+1, column=3).value.upper()
next_player=Label(root,text="NEXT:"+player_name2,fg="white",bg='black',
                  relief=RAISED,anchor="e",font=("Comic Sans MS",int(height/27)))
next_player.place(anchor=tk.SW, relx=0.72,rely=0.51)
current_weight_label=Label(root,text="CURRENT WEIGHT:",fg="white",bg='black',
                           relief=RAISED,anchor="e",font=("Comic Sans MS",30))
current_weight_label.place(anchor=tk.SW, relx=0.701,rely=0.60)
current_weight=StringVar()
current_weight_Entry= Entry(root, width=2, font=("Arial",150,""))
current_weight_Entry.insert(0, '20')
current_weight_Entry.place(x=(width-3*int(width/14)),y=height-4*int(height/10)-10,height=2*int(height/10))

pass_btn=Button(root,text="PASS",fg="white",bg="green",command=pass1)
pass_btn.place(x=(width-2.5*int(width/14)),y=height-2*int(height/10))
fail_btn=Button(root,text="FAIL",fg="white",bg="red",command=fail)
fail_btn.place(x=(width-3*int(width/14)),y=height-2*int(height/10))
submit(temp)
root.update()

#data saved

# infinite loop which is required to
# run tkinter program infinitely
# until an interrupt occurs


root.mainloop()
