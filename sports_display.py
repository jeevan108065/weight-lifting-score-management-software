
import time
from tkinter import *
from tkinter import messagebox
from tkinter import ttk, filedialog
import numpy
import pandas as pd
import pyautogui
from PIL import ImageTk, Image
from openpyxl import load_workbook
from openpyxl import *
from playsound import playsound

n=2
width, height= pyautogui.size()

# creating Tk window
root = Tk()

# setting geometry of tk window
root.state('zoomed')
root.configure(bg='black')
# Using title() to display a message in
# the dialogue box of the message in the
# title bar.
root.title("WEIGHT LIFTING")

#logo placing
img =Image.open('logopng.png')
bg = ImageTk.PhotoImage(img)
bg_image= Label(root, image=bg,bg='black')
bg_image.place(x=0,y=0)
img2 =Image.open('AIU_logo.png')
bg2 = ImageTk.PhotoImage(img2)
bg_image2= Label(root, image=bg2,bg='black')
bg_image2.place(x=width-1.2*width/15,y=0)
competition_name_label=Label(root,text="All India Inter-Univesity \n Weightlifting Women Championship 2021-22",
                             anchor="c",fg="red",bg='black',font=("Comic Sans MS",int(width*height/51840),"bold"))
competition_name_label.place(x=width-13.5*width/15,y=0)
#logo ends

#openinng players data

workbook = Workbook()
try:
    workbook = load_workbook(filename="sports_test_sorted.xlsx")
except:
    print("no file")
global sheet
sheet = workbook.active


#timer code
# Declaration of variables
hour=StringVar()
minute=StringVar()
second=StringVar()
# setting the default value as 0
hour.set("00")
minute.set("00")
second.set("60")
# Use of Entry class to take input from the user
hourEntry= Entry(root, width=3, font=("Arial",58,""),
				textvariable=hour)
minuteEntry= Entry(root, width=3, font=("Arial",58,""),
				textvariable=minute)
#minuteEntry.place(x=(width/2)-58,y=20)
secondEntry= Entry(root, width=3,fg="white",bg='black', font=("Arial",80,""),textvariable=second)
secondEntry.place(x=(width/2),y=height-(height/4))
Countdown_label=Label(root,text="COUNT DOWN:",anchor="e",font=("Comic Sans MS",30))
Countdown_label.place(x=4.2*width/15,y=height-(height/4)+20)

# the input provided by the user is
# stored in here :temp
global temp
global stop

temp = int(second.get())
def submit(temp,stop=False):
    from playsound import playsound
    #temp = int(second.get())
    # divmod(firstvalue = temp//60, secondvalue = temp%60)
    # mins,secs = divmod(temp,60)
	# Converting the input entered in mins or secs to hours,
	# mins ,secs(input = 110 min --> 120*60 = 6600 => 1hr :
	# 50min: 0sec)
    #hours=0
    secs=temp
    #if mins >60:		
		# divmod(firstvalue = temp//60, secondvalue
		# = temp%60)
        #hours, mins = divmod(mins, 60)		
	# using format () method to store the value up to
	# two decimal places
    #hour.set("{0:2d}".format(hours))
    #minute.set("{0:2d}".format(mins))
    second.set("{0:3d}".format(secs))
	# updating the GUI window after decrementing the
	# temp value every time
    root.update()
    time.sleep(1)
    # when temp value = 0; then a messagebox pop's up
	# with a message:"Time's up"
    if (temp<30):secondEntry['fg']="red"
    if (temp==30):
        try:
            playsound("beep-1.mp3")
        except:
            print("no audio")
    if (temp==90):
        try:
            playsound("beep-1.mp3")
        except:
            print("no audio")
    if (temp == 0):
        try:
            playsound("beep-2.mp3")
        except:
            print("no audio")
        messagebox.showinfo("Time Countdown", "Time's up ")
    # after every one sec the value of temp will be decremented
	# by one
    if stop==False:
        temp -= 1
    if temp >-1:
        submit(temp,stop)


def next_person():
    global n
    n=2
    r=2
    from openpyxl import load_workbook,Workbook
    workbook = Workbook()
    try:
        workbook = load_workbook(filename="sports_test.xlsx")
    except:
        print("no file")
    sheet = workbook.active
    for i in range(2,100):
        if sheet.cell(row=i,column=2).value == int(status_display.get()) :
            r=i
            break
        else :
            r=r+1
    save(r-1,trail1_Entry.get(),trail2_Entry.get(),trail3_Entry.get())
    sort()
def load():
    r=2
    workbook = Workbook()
    try:
        workbook = load_workbook(filename="sports_test_sorted.xlsx")
    except:
        print("no file")
    sheet = workbook.active
    n=int(status_display.get())+1 
    chest_no.set(sheet.cell(row=n, column=1).value)
    player_label['text'] =sheet.cell(row=n, column=2).value.upper()
    university_label['text'] =sheet.cell(row=n, column=3).value.upper()
    category_label['text'] ="CATEGORY:"+sheet.cell(row=n, column=4).value.upper()
    if event_name=="SNATCH":
        trail1.set(sheet.cell(row=n, column=5).value)
        trail2.set(sheet.cell(row=n, column=6).value)
        trail3.set(sheet.cell(row=n, column=7).value)
    elif event_name=="CLEAN_&_JERK":
        trail1.set(sheet.cell(row=n, column=9).value)
        trail2.set(sheet.cell(row=n, column=10).value)
        trail3.set(sheet.cell(row=n, column=11).value)
        
    secondEntry['fg']='white'
    if trail1.get()=="?":trail1_Entry['bg']="yellow"
    elif trail1.get()=="-":trail1_Entry['bg']="red"
    else:trail1_Entry['bg']="green"
    if trail2.get()=="?":trail2_Entry['bg']="yellow"
    elif trail2.get()=="-":trail2_Entry['bg']="red"
    else:trail2_Entry['bg']="green"
    if trail3.get()=="?":trail3_Entry['bg']="yellow"
    elif trail3.get()=="-":trail3_Entry['bg']="red"
    else:trail3_Entry['bg']="green"
    submit(temp)
    root.update()
    
def change_weight():
    next_weight_win=Tk()
    next_weight_win.geometry("100x100+"+str(width-250)+"+"+str(height-250))
    next_weight_win.configure(bg='yellow')
    next_weight_Entry=Entry(next_weight_win,width=3)
    next_weight_Entry.pack()
    def add():
        workbook = Workbook()
        try:
            workbook = load_workbook(filename="sports_test.xlsx")
        except:
            print("no file")
        sheet = workbook.active
        i=0
        for j in range(2,50):
            if sheet.cell(row=j, column=2).value==int(status_display.get()):
                i=j
                break
            else:
                i+=1
        cell1="R"+str(i)
        if trail3_Entry.get()=="?":
            sheet[cell1]=int(next_weight_Entry.get())
        else:
            sheet[cell1]='completed'
        workbook.save(filename="sports_test.xlsx")
        sort()
        next_btn['state']='active'
        next_weight_win.destroy()
    add_btn=Button(next_weight_win,text="ADD",command=add)
    add_btn.pack()
    next_weight_win.mainloop()
def sort():
    import pandas as pd
    xl = pd.ExcelFile("sports_test.xlsx")
    df = xl.parse("Sheet")
    df = df.sort_values(by=["trunk_no"])
    writer = pd.ExcelWriter('sports_test_sorted.xlsx')
    df.to_excel(writer,sheet_name='Sheet',columns=['trunk_no','name','university','category',
                                                   'attempt1','attempt2','attempt3','max1',
                                                   'attempt21','attempt22','attempt23','max2',
                                                   'total','next_weight'],
                index=False)
    writer.save()
    writer.close()	
sort()
#timer code ends
next_btn=Button(root,text="SAVE",command=next_person)
next_btn.place(x=(width-2*int(width/14)),y=height-2*int(height/10))
load_btn=Button(root,text="NEXT",command=load)
load_btn.place(x=20,y=height-1.5*int(height/10))
change_weight_btn=Button(root,text="CHANGE WEIGHT",command=change_weight)
#change_weight_btn.place(x=(width-2*int(width/14)),y=height-1.5*int(height/10))
#jersey display
chest_no=StringVar()
chest_label=Label(root,text="TRUNK NO:",anchor="e",fg="white",bg='black',font=("Comic Sans MS",30))
chest_label.place(x=10,y=3*height/10)
status_display= Entry(root, width=2,font=("Comic Sans MS",200),textvariable=chest_no)
status_display.insert(0,sheet.cell(row=n, column=1).value)
status_display.place(x=10,y=4*height/10)
#jersey display end

#player name
player_name=sheet.cell(row=n, column=2).value.upper()
player_label=Label(root,text=player_name,fg="white",bg='black', relief=RAISED,anchor="c",font=("Comic Sans MS",40))
player_label.place(x=7*width/15,y=2*height/10)
#player name end

#university name
university_name=sheet.cell(row=n, column=3).value.upper()
university_label=Label(root,text=university_name,fg="white",bg='black',anchor="e",font=("Comic Sans MS",30))
university_label.place(x=4.5*width/15,y=3*height/10)
#university name ends

#category
category_name=sheet.cell(row=n, column=4).value.upper()
category_label=Label(root,text="CATEGORY:"+category_name,fg="yellow", bg='black',relief=RAISED,anchor="e",font=("Comic Sans MS",30))
category_label.place(x=10,y=2*height/10)
#category end

#event name
event_name="SNATCH"
event_label=Label(root,text=event_name,fg="white",bg='black', relief=SUNKEN,anchor="e",font=("Comic Sans MS",30))
event_label.place(x=7.2*width/15,y=4*height/10)
#event name end
def cEvent():
    global event_name
    event_name="CLEAN_&_JERK"
    event_label['text']=event_name
    event_label.place(x=6.5*width/15,y=4*height/10)
    trail1_Entry.delete(0, 'end')
    trail2_Entry.delete(0, 'end')
    trail3_Entry.delete(0, 'end')
    trail1_Entry['bg']='yellow'
    trail2_Entry['bg']='yellow'
    trail3_Entry['bg']='yellow'
    trail1_Entry.insert(0,sheet.cell(row=n, column=9).value)
    trail2_Entry.insert(0,sheet.cell(row=n, column=10).value)
    trail3_Entry.insert(0,sheet.cell(row=n, column=11).value)
    secondEntry['fg']='white'
    if trail1.get()=="?":trail1_Entry['bg']="yellow"
    elif trail1.get()=="-":trail1_Entry['bg']="red"
    else:trail1_Entry['bg']="green"
    if trail2.get()=="?":trail2_Entry['bg']="yellow"
    elif trail2.get()=="-":trail2_Entry['bg']="red"
    else:trail2_Entry['bg']="green"
    if trail3.get()=="?":trail3_Entry['bg']="yellow"
    elif trail3.get()=="-":trail3_Entry['bg']="red"
    else:trail3_Entry['bg']="green"
#score card
trail1=StringVar()
trail2=StringVar()
trail3=StringVar()

if event_name=="SNATCH":
    trail1_Entry= Entry(root, width=3, font=("Arial",58,""),textvariable=trail1)
    trail1_Entry.insert(0,sheet.cell(row=n, column=5).value)
    if trail1.get()=="?":trail1_Entry['bg']="yellow"
    elif trail1.get()=="-":trail1_Entry['bg']="red"
    else:trail1_Entry['bg']="green"
    trail2_Entry= Entry(root, width=3, font=("Arial",58,""),textvariable=trail2)
    trail2_Entry.insert(0,sheet.cell(row=n, column=6).value)
    if trail2.get()=="?":trail2_Entry['bg']="yellow"
    elif trail2.get()=="-":trail2_Entry['bg']="red"
    else:trail2_Entry['bg']="green"
    trail3_Entry= Entry(root, width=3, font=("Arial",58,""),textvariable=trail3)
    trail3_Entry.insert(0,sheet.cell(row=n, column=7).value)
    if trail3.get()=="?":trail3_Entry['bg']="yellow"
    elif trail3.get()=="-":trail3_Entry['bg']="red"
    else:trail3_Entry['bg']="green"
elif event_name=="CLEAN_&_JERK":
    trail1_Entry= Entry(root, width=3, font=("Arial",58,""),textvariable=trail1)
    trail1_Entry.insert(0,sheet.cell(row=n, column=11).value)
    if trail1.get()=="?":trail1_Entry['bg']="yellow"
    elif trail1.get()=="-":trail1_Entry['bg']="red"
    else:trail1_Entry['bg']="green"
    trail2_Entry= Entry(root, width=3, font=("Arial",58,""),textvariable=trail2)
    trail2_Entry.insert(0,sheet.cell(row=n, column=12).value)
    if trail2.get()=="?":trail2_Entry['bg']="yellow"
    elif trail2.get()=="-":trail2_Entry['bg']="red"
    else:trail2_Entry['bg']="green"
    trail3_Entry= Entry(root, width=3, font=("Arial",58,""),textvariable=trail3)
    trail3_Entry.insert(0,sheet.cell(row=n, column=13).value)
    if trail3.get()=="?":trail3_Entry['bg']="yellow"
    elif trail3.get()=="-":trail3_Entry['bg']="red"
    else:trail3_Entry['bg']="green"

trail1_Entry.place(x=6*width/15,y=5*height/10, height=200)
trail2_Entry.place(x=7.4*width/15,y=5*height/10, height=200)
trail3_Entry.place(x=8.8*width/15,y=5*height/10, height=200)
#score card end

#adding data to excel
cell1="E"+str(n)
cell2="F"+str(n)
cell3="G"+str(n)
sheet[cell1]=trail1_Entry.get()
sheet[cell2]=trail2_Entry.get()
sheet[cell3]=trail3_Entry.get()
def save(n,t1,t2,t3):
    from openpyxl import load_workbook,Workbook
    workbook = Workbook()
    try:
        workbook = load_workbook(filename="sports_test.xlsx")
    except:
        print("no file")
    sheet = workbook.active
    if event_name=="SNATCH":
        cell1="F"+str(n+1)
        cell2="G"+str(n+1)
        cell3="H"+str(n+1)
        cell4="I"+str(n+1)
        cell5="J"+str(n+1)
    elif event_name=="CLEAN_&_JERK":
        cell1="K"+str(n+1)
        cell2="L"+str(n+1)
        cell3="M"+str(n+1)
        cell4="N"+str(n+1)
        cell5="O"+str(n+1)
    if trail1_Entry.get()!="?" and trail1_Entry.get()!="-":
        sheet[cell1]=int(trail1_Entry.get())
    else:
        sheet[cell1]=trail1_Entry.get()
    if trail2_Entry.get()!="?" and trail2_Entry.get()!="-":
        sheet[cell2]=int(trail2_Entry.get())
    else:
        sheet[cell2]=trail2_Entry.get()
    if trail3_Entry.get()!="?" and trail3_Entry.get()!="-":
        sheet[cell3]=int(trail3_Entry.get())
    else:
        sheet[cell3]=trail3_Entry.get()
    ma=0
    if t1=="?":
        jk=0
        t1=0
    elif t2=="?":
        jk=1
        t2=0
    elif t3=="?":
        jk=2
        t3=0
    else:
        jk=3
    we=[]
    if str(t1).isnumeric():
        we.append(t1)
    elif str(t2).isnumeric():
        we.append(t2)
    elif str(t3).isnumeric():
        we.append(t3)
    try:
        ma=max(we)
    except:
        ma=0
    sheet[cell4]=int(ma)
    sheet[cell5]=jk
    cell6="P"+str(n+1)
    sheet[cell6]=int(sheet.cell(row=n+1, column=9).value)+int(sheet.cell(row=n+1, column=14).value)
    workbook.save(filename="sports_test.xlsx")
save(n,trail1_Entry.get(),trail2_Entry.get(),trail3_Entry.get())
def pass1():
    from playsound import playsound
    if trail1_Entry.get()=="?" and trail1_Entry.get()!="-":
        trail1_Entry.delete(0, 'end')
        trail1_Entry.insert(0,current_weight_Entry.get())
        trail1_Entry['bg']="green"
    elif trail2_Entry.get()=="?" and trail2_Entry.get()!="-":
        trail2_Entry.delete(0, 'end')
        trail2_Entry.insert(0,current_weight_Entry.get())
        trail2_Entry['bg']="green"
    elif trail3_Entry.get()=="?" and trail3_Entry.get()!="-":
        trail3_Entry.delete(0, 'end')
        trail3_Entry.insert(0,current_weight_Entry.get())
        trail3_Entry['bg']="green"
    stop=True
    try:
        playsound("beep-1.mp3")
    except:
        print("no audio")
    submit(temp,stop)
    root.update()
    pass
def fail():
    from playsound import playsound
    if trail1_Entry.get()=="?" and trail1_Entry.get()!="-":
        trail1_Entry.delete(0, 'end')
        trail1_Entry.insert(0,"-")
        trail1_Entry['bg']="red"
    elif trail2_Entry.get()=="?" and trail2_Entry.get()!="-":
        trail2_Entry.delete(0, 'end')
        trail2_Entry.insert(0,"-")
        trail2_Entry['bg']="red"
    elif trail3_Entry.get()=="?" and trail3_Entry.get()!="-":
        trail3_Entry.delete(0, 'end')
        trail3_Entry.insert(0,"-")
        trail3_Entry['bg']="red"
    stop=True
    try:
        playsound("beep-1.mp3")
    except:
        print("no audio")
    submit(temp,stop)
    root.update()
    pass
#current weight
player_name2=sheet.cell(row=n+1, column=2).value.upper()
next_player=Label(root,text="NEXT:"+player_name2,fg="white",bg='black',
                  relief=RAISED,anchor="e",font=("Comic Sans MS",30))
#next_player.place(x=(width-3.8*int(width/14)),y=height-5.6*int(height/10)-10)
current_weight_label=Label(root,text="CURRENT WEIGHT:",fg="white",bg='black',
                           relief=RAISED,anchor="e",font=("Comic Sans MS",30))
current_weight_label.place(x=(width-3.7*int(width/14)),y=height-4.8*int(height/10)-10)
current_weight=StringVar()
current_weight_Entry= Entry(root, width=3, font=("Arial",150,""))
current_weight_Entry.insert(0, '15')
current_weight_Entry.place(x=(width-3.7*int(width/14)),y=height-4*int(height/10)-10,height=2*int(height/10))

pass_btn=Button(root,text="PASS",fg="white",bg="green",command=pass1)
pass_btn.place(x=(width-2.5*int(width/14)),y=height-2*int(height/10))
fail_btn=Button(root,text="FAIL",fg="white",bg="red",command=fail)
fail_btn.place(x=(width-3*int(width/14)),y=height-2*int(height/10))
cEvent_btn=Button(root,text="NEXT EVENT",command=cEvent)
cEvent_btn.place(x=(width-3*int(width/14)),y=height-1.5*int(height/10))
submit(temp)
root.update()

#data saved

# infinite loop which is required to
# run tkinter program infinitely
# until an interrupt occurs


root.mainloop()
