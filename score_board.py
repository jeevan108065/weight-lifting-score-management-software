
# Python program to create a table
import time   
from tkinter import *
from PIL import ImageTk, Image
from tkscrolledframe import ScrolledFrame
import pyautogui
width, height= pyautogui.size()
  
class Table:
      
    def __init__(self,root,c,d):
        #logo placing
        self.img =Image.open('logopng.png')
        self.bg = ImageTk.PhotoImage(self.img)
        self.bg_image= Label(root, image=self.bg,bg='black')
        self.bg_image.grid(row=0,column=0)
        self.img2 =Image.open('AIU_logo.png')
        self.bg2 = ImageTk.PhotoImage(self.img2)
        self.bg_image2= Label(root, image=self.bg2,bg='black')
        self.bg_image2.place(x=width-1.5*width/15,y=0)
        self.competition_name_label=Label(root,
                                          text="All India Inter-Univesity \n Weightlifting Women Championship 2021-22 \n Category:81KgS",
                                          anchor="c",fg="white",bg='black',font=("Times New Roman",40,"bold"))
        self.competition_name_label.place(x=width-13*width/15,y=0)
        #logo ends
        for i in range(9):
            self.e = Entry(root, width=0,bd=0,bg='black')
            self.e.grid(row=i, column=1)
        # code for creating table
        self.e = Label(root, text="Snatch",fg='white',bg='black',font=('Arial',11,'bold'))
        self.e.grid(row=10, column=5)
        self.e = Label(root,text="Clean&Jerk",fg='white',bg='black',font=('Arial',11,'bold'))
        self.e.grid(row=10, column=8)
        for i in range(total_rows):
            for j in range(total_columns):
                if j==3:
                    continue
                val=""
                if lst[i][j] is not None:val=lst[i][j]
                if j==8 or j==12:
                    if i==0:
                        self.e = Label(root,text=val,width=10,border=5,anchor='w',
                                       bd=2,fg='white',bg='black',font=('Arial',12,'bold'))
                    elif i==1:
                        self.e = Label(root,text=val,width=10,border=5,anchor='w',
                                       bd=2, fg='black',bg='#90EE90',font=('Arial',12,'bold'))
                    elif i==2:
                        self.e = Label(root,text=val,width=10,border=5,anchor='w',
                                       bd=2, fg=d,bg='orange',font=('Arial',12,'bold'))
                    elif i==3:
                        self.e = Label(root,text=val,width=10,border=5,anchor='w',
                                       bd=2, fg='black',bg='#FFFF99',font=('Arial',12,'bold'))
                    else:
                        self.e = Label(root,text=val,width=10,border=5,anchor='w',
                                       bd=2, fg='blue',font=('Arial',12,'bold'))
                else:
                    if i==0:
                        self.e = Label(root,text=val,width=8,border=5,anchor='w',
                                       bd=2,fg='white',bg='black',font=('Arial',12,'bold'))
                    elif i==1:
                        self.e = Label(root,text=val,width=8,border=5,anchor='w',
                                       bd=2, fg='black',bg='#90EE90',font=('Arial',12,'bold'))
                    elif i==2:
                        self.e = Label(root,text=val,width=8,border=5,anchor='w',
                                       bd=2, fg=d,bg='orange',font=('Arial',12,'bold'))
                    elif i==3:
                        self.e = Label(root,text=val,width=8,border=5,anchor='w',
                                       bd=2, fg='black',bg='#FFFF99',font=('Arial',12,'bold'))
                    else:
                        self.e = Label(root,text=val,width=8,border=5,anchor='w',
                                       bd=2, fg='blue',font=('Arial',12,'bold'))
                if j==1:self.e['width']=25
                if j==2:self.e['width']=45
                self.e.grid(row=i+12, column=j)
                if lst[i][12]==0:
                    self.e['bg']='red'
                if (j>3 and j<7) or (j>7 and j<11):
                    if lst[i][j]=="-":
                        self.e['bg']='red'
                    elif str(lst[i][j]).isnumeric():
                        self.e['bg']='#00FF00'
        root.update()
        
        
root = Tk()
root.config(bg='black')
root.state('zoomed')
from openpyxl import load_workbook
from openpyxl import *
workbook = Workbook()
try:
    workbook = load_workbook(filename="sports_test_sorted.xlsx")
except:
    print("no file")
global sheet
sheet = workbook.active
global lst
lst = []
c=[]
r=4
for i in range(4,100):
    if sheet.cell(row=i,column=1).value is None :
        break
    else :
        r=r+1
for j in range(1,r+1):
    c=[]
    for i in range(1,14):
        c.append(sheet.cell(row=j, column=i).value)
    lst.append(c)
# find total number of rows and
# columns in list
global total_rows
global total_columns
total_rows = len(lst)-1
total_columns = len(lst[0])  
t = Table(root,"orange",'black')
def refresh():
    root.destroy()
    import os
    os.system('python score_board.py')
c=0
r=0    
def countdown(c):
    global r
    root.update()
    time.sleep(1)
    r+=1
    if r==15:
        refresh()
    countdown(c^1)
countdown(c)
workbook.close()
root.mainloop()
