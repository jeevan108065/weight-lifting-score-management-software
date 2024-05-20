import time   
from tkinter import *
from PIL import ImageTk, Image
from tkscrolledframe import ScrolledFrame
import pyautogui
import os
width, height= pyautogui.size()
root = Tk()
root.config(bg='black')
root.state('zoomed')
from openpyxl import load_workbook
from openpyxl import *
workbook = Workbook()
try:
    workbook = load_workbook(filename="sports_test_winners.xlsx")
except:
    print("no file")
global sheet
sheet = workbook.active
global lst
lst = []
c=[]
r=0
for i in range(1,100):
    if sheet.cell(row=i,column=1).value is None :
        break
    else :
        r=r+1
for j in range(1,r+1):
    c=[]
    for i in range(1,8):
        c.append(sheet.cell(row=j, column=i).value)
    lst.append(c)
root.img =Image.open('logopng.png')
root.bg = ImageTk.PhotoImage(root.img)
root.bg_image= Label(root, image=root.bg,bg='black')
root.bg_image.place(x=70,y=0)
root.img2 =Image.open('AIU_logo.png')
root.bg2 = ImageTk.PhotoImage(root.img2)
root.bg_image2= Label(root, image=root.bg2,bg='black')
root.bg_image2.place(x=width-1.5*width/15,y=0)
root.competition_name_label=Label(root,
                                  text="All India Inter-Univesity \n Weightlifting Women Championship 2021-22",
                                  anchor="c",fg="white",bg='black',font=("Times New Roman",40,"bold"))
root.competition_name_label.place(x=width-13*width/15,y=0)
n=0
root.category= Label(root, text=lst[n][0],fg='lightblue',bg='black',font=('Arial',60,'bold'))
root.category.place(x=700,y=200)
root.place= Label(root, text=lst[n][1],fg='black',bg='white',font=('Arial',60,'bold'))
root.place.place(x=300,y=200)
root.name= Label(root, text=lst[n][2],fg='lightblue',bg='black',font=('Arial',100,'bold'))
root.name.place(x=300,y=400)
root.university= Label(root, text=lst[n][3],fg='lightblue',bg='black',font=('Arial',60,'bold'))
root.university.place(x=300,y=600)
root.snatch= Label(root, text=lst[n][4],fg='lightblue',bg='black',font=('Arial',60,'bold'))
root.snatch.place(x=300,y=700)
root.clean= Label(root, text=lst[n][5],fg='lightblue',bg='black',font=('Arial',60,'bold'))
root.clean.place(x=300,y=800)
root.total= Label(root, text=lst[n][6],fg='lightblue',bg='black',font=('Arial',60,'bold'))
root.total.place(x=300,y=900)
def next_name():
    global n
    if n==len(lst)-1:
        root.destroy()
        os.system('python loop.py')
    n=n+1
    root.category['text']=lst[n][0]
    root.place['text']=lst[n][1]
    root.name['text']=lst[n][2]
    root.university['text']=lst[n][3]
    root.snatch['text']=lst[n][4]
    root.clean['text']=lst[n][5]
    root.total['text']=lst[n][6]
    if n==len(lst)-1:
        n==0
    root.update()
    time.sleep(5)
    next_name()
def before_name():
    global n
    if n==0:
        n==len(lst)-1
    n=n-1
    root.category['text']=lst[n][0]
    root.name['text']=lst[n][1]
    root.university['text']=lst[n][2]
    root.snatch['text']=lst[n][3]
    root.clean['text']=lst[n][4]
    root.total['text']=lst[n][5]
    root.update()
B_btn=Button(root,text="BACK",command=before_name)
B_btn.place(x=100,y=height-100)
btn=Button(root,text="NEXT",command=next_name)
btn.place(x=width-100,y=height-100)
next_name()
root.mainloop()